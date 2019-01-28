#-*- coding: utf-8 -*-
# @Time   : 2019/1/10 11:04
# @Author : linglong
# @File   : myExcel.py
from copy import copy

import openpyxl as openpyxl
import xlrd
import xlwt as xlwt
import time


class NewExcel(object):
    """
    操作excel表格类
    """
    def __init__(self,path):
        """
        初始化赋值
        :param path:
        """
        self.path = path

    def read_excel(self):
        """
        读取表的所有值
        :return:返回所有页签的所有值  result[x][y][z]  x-页签名称 y-1 第二行  z-2 第三列
        """
        try:
            workbook = xlrd.open_workbook(self.path)  # 打开工作簿
            sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        except BaseException as e:
            return str(e)

        #循环遍历,把所有表的数据返回
        result = {}
        for x in range(len(sheets)):
            value = {}
            worksheet = workbook.sheet_by_name(sheets[x])
            for i in range(0, worksheet.nrows):
                l = []
                for j in range(0, worksheet.ncols):
                    l.append(worksheet.cell_value(i, j))
                value[i] = l
            result[workbook.sheet_names()[x]] = value
        return result

    def create_excel(self,value,book_name=str(time.time())[0:9]+'.xlsx'):
        """
        创建新的表格。目前没有应用创建，所以暂时实行创建一个页签的表格
        :param value: 表格的值 格式：[[],[],[]]
        :param book_name: 文件的保存路径
        :return:
        """
        try:
            index = len(value)  # 获取需要写入数据的行数
            workbook = xlwt.Workbook()  # 新建一个工作簿
            sheet = workbook.add_sheet('autoCreateSheet')  # 在工作簿中新建一个表格
            for i in range(0, index):
                for j in range(0, len(value[i])):
                    sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
            workbook.save(book_name)  # 保存工作簿
        except BaseException as e:
            return str(e)
        return

    def edit_excel(self,value):
        """
        编辑存在的文件,因为没有具体的应用场景,现在实行的就是传入的新值覆盖旧值
        :param value: 新数据的值
        :return:
        """
        try:
            index = len(value)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = '重新写入了数据'
            for i in range(0, index):
                for j in range(0, len(value[i])):
                    sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
            workbook.save(self.path)
        except BaseException as e:
            return str(e)
        return



#
# value = [["张三", "男", "19", "杭州", "研发工程师"],
#           ["李四", "男", "22", "北京", "医生"],
#           ["王五", "女", "33", "珠海", "出租车司机"],]
#
# value1 = [["张三1", "男", "19", "杭州", "研发工程师"],
#           ["李四2", "男", "22", "北京", "医生"],
#           ["王五3", "女", "33", "珠海", "出租车司机"],]
#
#
# t = NewExcel("test1.xlsx")
# t.create_excel(value,"test1.xlsx")
# print(t.read_excel())
# time.sleep(60)
# t.edit_excel(value1)
# print(t.read_excel())
